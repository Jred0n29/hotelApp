from django.shortcuts import render, redirect
from .models import Clientes
from datetime import date,datetime
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse_lazy
# Create your views here.
from django.views.generic import TemplateView
from openpyxl import Workbook
from django.db.models import Sum
from django.db.models.functions import Coalesce
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa


def home(request, *args, **kwargs):
    """Inicio de la aplicación"""
    clientes = Clientes.objects.all()

    ingresos_mensuales = 0
    proximos_inactivos = 0
    nuevos_usuarios = 0
    ingresos_diarios = 0
    personas_inactivas = []
    data = []
    try:
        year = datetime.now().year
        for m in range(1, 13):
            total = Clientes.objects.filter(fecha_inicial__year=year, fecha_inicial__month=m).aggregate(r=Coalesce(Sum('valor'), 0)).get('r')
            data.append(float(total))
    except:
        pass
    
            
    for i in clientes:

        dias_restantes = i.fecha_final - i.fecha_inicial

        if i.fecha_inicial.month == date.today().month:

            """ Visualizar ingresos mensuales y usuarios mensuales"""
            ingresos_mensuales = i.valor + ingresos_mensuales
            nuevos_usuarios = nuevos_usuarios+1

        if dias_restantes.days <= 7 and date.today() <= i.fecha_final:
            """ Visualizar los proximos inactivos """
            proximos_inactivos = proximos_inactivos+1
            personas_inactivas.append(f"""{i.nombres} {i.apellidos} """)


        if i.fecha_final <= date.today():
            """ Cambiar el estado cuando se haya vencido su instancia"""
            i.estado = False
            i.save()

        if i.fecha_inicial == date.today():
            """Ingresos Diarios"""
            ingresos_diarios = i.valor + ingresos_diarios


    return render(request, "index.html", {
        'clientes': clientes,
                                        'ingresos_mensuales': ingresos_mensuales,
                                        'proximos_inactivos': proximos_inactivos,
                                        'nuevos_usuarios': nuevos_usuarios,
                                        'ingresos_diarios': ingresos_diarios,
                                        'personas_inactivas': personas_inactivas,
                                        'estadisticas':data
        })


def registrar_clientes(request):
    modelos = [
        'identificacion',
                'nombres',
                'apellidos',
                'telefono',
                'fecha_inicial',
                'fecha_final',
                'valor',
                'comentario'
        ]
    datos_clientes = []
    if request.method == 'POST':
        for i in modelos:
            lista_datos = request.POST.get(i)
            datos_clientes.append(lista_datos)

        nuevos_datos = Clientes(identificacion= datos_clientes[0],
                                nombres= datos_clientes[1],
                                apellidos= datos_clientes[2],
                                telefono= datos_clientes[3],
                                fecha_inicial= datos_clientes[4],
                                fecha_final= datos_clientes[5],
                                valor= datos_clientes[6],
                                comentario= datos_clientes[7])


        print(nuevos_datos)
        nuevos_datos.save()
        return redirect('index')

    return render(request, "clientes.html")


def baseD(request):
    """Inicio de la aplicación"""
    clientes = Clientes.objects.all()
    return render(request, "bd.html", {'clientes': clientes})

def eliminarC(request, id):
    cliente = Clientes.objects.get(id= id)
    cliente.delete()
    return redirect('hotel:datos')

def ocultarC(request, id):
    cliente = Clientes.objects.get(id= id)
    cliente.ocultar = False
    cliente.save()
    return redirect('index')

def facturasC(request, id):
    """Facturas de los clientes"""
    try:
        cliente = Clientes.objects.get(id= id)
        fecha_actual = date.today()
        template = get_template('facturas.html')
        context = {'cliente': cliente, 'fecha_actual':fecha_actual}
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
           #response['Content-Disposition'] = 'attachment; filename="report.pdf"'
        pisaStatus = pisa.CreatePDF(
            html, dest=response)
        return response
    except:
        pass
    return HttpResponseRedirect(reverse_lazy('index'))


def generarExcel(resquest):
    clientes = Clientes.objects.all()
    wb = Workbook()
    bandera = True
    cont = 1
    controlador = 4
    for i in clientes:
        if bandera:
            ws = wb.active
            ws.title = 'Hoja'+str(cont)
            bandera = False

        # Crear el título en la hoja
        ws['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B1'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B1'].fill = PatternFill(start_color= '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['B1'].font = Font(name= 'Calibri', size = 12, bold = True)
        ws['B1'] = 'REPORTE PERSONALIZADO EN EXCEL CON DJANGO'

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:G1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

        # Crear la cabecera
        ws['B3'].alignment = Alignment(horizontal= "center", vertical = "center")
        ws['B3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"),
                                 top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B3'].fill = PatternFill(start_color= '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B3'].font = Font(name= 'Calibro', size = 10, bold = True)
        ws['B3'] = 'Nombres'

        ws['C3'].alignment = Alignment(horizontal= "center", vertical = "center")
        ws['C3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"),
                                 top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['C3'].fill = PatternFill(start_color= '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['C3'].font = Font(name= 'Calibro', size = 10, bold = True)
        ws['C3'] = 'Apellidos'

        ws['D3'].alignment = Alignment(horizontal= "center", vertical = "center")
        ws['D3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"),
                                 top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['D3'].fill = PatternFill(start_color= '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['D3'].font = Font(name= 'Calibro', size = 10, bold = True)
        ws['D3'] = 'Ideintificacion'

        ws['E3'].alignment = Alignment(horizontal= "center", vertical = "center")
        ws['E3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"),
                                 top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E3'].fill = PatternFill(start_color= '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['E3'].font = Font(name= 'Calibro', size = 10, bold = True)
        ws['E3'] = 'Estado'

        ws['F3'].alignment = Alignment(horizontal= "center", vertical = "center")
        ws['F3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"),
                                 top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['F3'].fill = PatternFill(start_color= '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['F3'].font = Font(name= 'Calibro', size = 10, bold = True)
        ws['F3'] = 'Fecha Inicio '

        ws['G3'].alignment = Alignment(horizontal= "center", vertical = "center")
        ws['G3'].border = Border(left= Side(border_style = "thin"), right = Side(border_style = "thin"),
                                 top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['G3'].fill = PatternFill(start_color= '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['G3'].font = Font(name= 'Calibro', size = 10, bold = True)
        ws['G3'] = 'Fecha Final'

        ws.cell(row= controlador, column = 2).alignment = Alignment(horizontal = "center")
        ws.cell(row= controlador, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                               top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row= controlador, column = 2).font = Font(name = 'Calibri', size = 8)
        ws.cell(row= controlador, column = 2).value = i.nombres

        ws.cell(row= controlador, column = 3).alignment = Alignment(horizontal = "center")
        ws.cell(row= controlador, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                               top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row= controlador, column = 3).font = Font(name = 'Calibri', size = 8)
        ws.cell(row= controlador, column = 3).value = i.apellidos

        ws.cell(row= controlador, column = 4).alignment = Alignment(horizontal = "center")
        ws.cell(row= controlador, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                               top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row= controlador, column = 4).font = Font(name = 'Calibri', size = 8)
        ws.cell(row= controlador, column = 4).value = i.identificacion

        ws.cell(row= controlador, column = 5).alignment = Alignment(horizontal = "center")
        ws.cell(row= controlador, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                               top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row= controlador, column = 5).font = Font(name = 'Calibri', size = 8)
        ws.cell(row= controlador, column = 5).value = i.estado

        ws.cell(row= controlador, column = 6).alignment = Alignment(horizontal = "center")
        ws.cell(row= controlador, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                               top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row= controlador, column = 6).font = Font(name = 'Calibri', size = 8)
        ws.cell(row= controlador, column = 6).value = i.fecha_inicial

        ws.cell(row= controlador, column = 7).alignment = Alignment(horizontal = "center")
        ws.cell(row= controlador, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                               top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws.cell(row= controlador, column = 7).font = Font(name = 'Calibri', size = 8)
        ws.cell(row= controlador, column = 7).value = i.fecha_final

        controlador = controlador+1

    # Establecer el nombre de mi archivo
    nombre_archivo = "ReportePersonalizadoExcel.xlsx"
    # Definir el tipo de respuesta que se va a dar
    response = HttpResponse(content_type= "application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


def editarC(request, id):
    try:
        cliente = Clientes.objects.get(id= id)
        modelos = [
            'identificacion',
                    'nombres',
                    'apellidos',
                    'telefono',
                    'fecha_inicial',
                    'fecha_final',
                    'valor',
                    'comentario'
            ]
        datos_clientes = []
        if request.method == 'POST':
            for i in modelos:
                lista_datos = request.POST.get(i)
                datos_clientes.append(lista_datos)

            nuevos_datos = Clientes(identificacion= datos_clientes[0],
                                    nombres= datos_clientes[1],
                                    apellidos= datos_clientes[2],
                                    telefono= datos_clientes[3],
                                    fecha_inicial= datos_clientes[4],
                                    fecha_final= datos_clientes[5],
                                    valor= datos_clientes[6],
                                    comentario= datos_clientes[7])

            nuevos_datos.save()
            cliente.delete()

            return redirect('hotel:datos')
        else:

            return render(request, "editar.html", {'cliente': cliente})

    except:
        return redirect('hotel:index')


class error404(TemplateView):
    """Error de la aplicación"""
    template_name = "404.html"


class error500(TemplateView):
    """Error de la aplicación"""
    template_name = "500.html"

    @classmethod
    def as_error_view(cls):

        v = cls.as_view()

        def view(request):
            r = v(request)
            r.render()
            return r
        return view

# Create your views here.
